"""Generate the Excel area breakdown for a job.

Single responsibility: produce an openpyxl workbook with one row per plot,
showing the survey-wise area breakdown. Generated in memory, returns bytes.

Two sheets:
- "Area Summary": one row per plot, key columns the client needs for their
  land-parcel-system workflow (survey no, location, areas, status, flags count).
- "Measurements": one row per accepted measurement across all plots, for
  the client's quality-control review.

Column widths and header formatting follow the client's existing Excel convention
(bold header row, frozen panes at row 2, auto-filtered, numeric area columns
formatted to 3 decimal places).
"""

from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ...core.enums import PlotStatus
from ...core.models import Job

__all__ = ["generate_excel_sheet"]

# Header fill colors matching the client's existing convention.
_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
_FLAGGED_FILL = PatternFill("solid", fgColor="FFF2CC")
_FAILED_FILL = PatternFill("solid", fgColor="FFE0E0")


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, values: list[str], row: int = 1) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def generate_excel_sheet(job: Job) -> bytes:
    """Generate the Excel area sheet for ``job`` and return its bytes.

    Returns:
        Raw .xlsx bytes ready to write or upload.
    """
    wb = openpyxl.Workbook()

    # --- Sheet 1: Area Summary ---
    ws1 = wb.active
    ws1.title = "Area Summary"

    headers = [
        "Survey No.", "District", "Taluk", "Village",
        "Stated Area (Ha)", "Computed Area (Ha)", "Error (%)",
        "Boundary", "Corner Stones", "Status", "Flags Count",
    ]
    _header_row(ws1, headers, row=1)
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    _set_col_widths(ws1, [12, 14, 14, 16, 16, 18, 12, 10, 14, 12, 12])

    for row_idx, plot in enumerate(job.plots, start=2):
        computed_ha = (
            round(plot.boundary.computed_area / 10_000.0, 3)
            if plot.boundary else None
        )
        stated = round(plot.stated_area, 3) if plot.stated_area is not None else None
        error_pct = None
        if stated and computed_ha is not None and stated > 0:
            error_pct = round(abs(computed_ha - stated) / stated * 100, 2)
        closed = (
            "Closed" if (plot.boundary and plot.boundary.is_closed)
            else "Open" if plot.boundary else None
        )
        values = [
            plot.survey_no,
            plot.district,
            plot.taluk,
            plot.village,
            stated,
            computed_ha,
            error_pct,
            closed,
            len(plot.corner_points),
            plot.status.value.replace("_", " ").title(),
            len(plot.flags),
        ]
        fill = (
            _FAILED_FILL if plot.status is PlotStatus.FAILED
            else _FLAGGED_FILL if plot.status is PlotStatus.FLAGGED
            else None
        )
        for col_idx, val in enumerate(values, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            if fill:
                cell.fill = fill
            # Format area columns (5 = Stated, 6 = Computed) as 3dp numbers.
            if col_idx in (5, 6) and val is not None:
                cell.number_format = "0.000"
            if col_idx == 7 and val is not None:
                cell.number_format = "0.00"

    # --- Sheet 2: Measurements ---
    ws2 = wb.create_sheet("Measurements")
    meas_headers = [
        "Survey No.", "Raw OCR", "Normalized Value", "Confidence",
        "Line Class", "Line Ref", "Source",
    ]
    _header_row(ws2, meas_headers, row=1)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(meas_headers))}1"
    _set_col_widths(ws2, [12, 14, 16, 12, 12, 18, 12])

    meas_row = 2
    for plot in job.plots:
        for m in plot.measurements:
            if m.value is None:
                continue  # only accepted measurements in the detail sheet
            ws2.cell(row=meas_row, column=1, value=plot.survey_no)
            ws2.cell(row=meas_row, column=2, value=m.raw)
            cell_val = ws2.cell(row=meas_row, column=3, value=m.value)
            cell_val.number_format = "0.000"
            ws2.cell(row=meas_row, column=4, value=round(m.confidence, 2))
            ws2.cell(row=meas_row, column=5, value=m.line_class)
            ws2.cell(row=meas_row, column=6, value=m.line_ref)
            ws2.cell(row=meas_row, column=7, value=m.source.value)
            meas_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
