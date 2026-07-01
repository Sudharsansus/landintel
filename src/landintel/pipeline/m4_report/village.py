"""M4 village deliverable -- the report a SPAN's M2/M3 output ships as.

The existing area_statement / excel_sheet generate per-JOB reports keyed off the
``Job``/``Plot`` model (the API path). M2's corridor output is a list of
``GeorefResult`` + one combined village DXF, with each FMB dispositioned
ACCEPT / REVIEW / NO_COVERAGE. This module turns THAT directly into a client
deliverable, so a corridor run produces a shippable package without going through
the API/Job model:

  * a village area-statement PDF (one row per survey: disposition, ground area,
    chain coverage, field residual) + a disposition summary;
  * a village Excel with the same rows (filterable);
  * a delivery .zip bundling the combined DXF + the two reports.

Areas come from each ACCEPT plot's georeferenced boundary (m3_assemble.area), so
the number on the report matches the verified geometry. Pure/in-memory generators
(return bytes) so they are testable without S3; ``build_village_delivery`` writes
the package locally and returns its path.
"""

from __future__ import annotations

import io
import logging
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

_log = logging.getLogger(__name__)

__all__ = [
    "VillageRow",
    "rows_from_results",
    "village_area_statement_pdf",
    "village_excel",
    "build_village_delivery",
]

_ACCEPT = ("ACCEPT", "ACCEPT_SEEDED", "ACCEPT_CADASTRAL")


@dataclass
class VillageRow:
    """One survey's line in the village deliverable."""
    survey_number: str
    recommendation: str
    area_ha: float | None
    chain_coverage: float
    field_residual_max: float
    match_method: str = ""

    @property
    def georeferenced(self) -> bool:
        return self.recommendation in _ACCEPT


def rows_from_results(results: Iterable[Any]) -> list[VillageRow]:
    """Build report rows from M2 ``GeorefResult`` objects (duck-typed).

    Ground area is read from each placed plot's georeferenced DXF boundary; a plot
    that was not georeferenced (REVIEW / NO_COVERAGE) has ``area_ha = None`` rather
    than a fabricated number."""
    from ..m3_assemble.area import plot_area_hectares

    rows: list[VillageRow] = []
    for r in results:
        area = None
        if r.recommendation in _ACCEPT and getattr(r, "output_file", ""):
            try:
                area = plot_area_hectares(r.output_file)
            except Exception:  # noqa: BLE001 - a bad file must not break the report
                area = None
        rows.append(VillageRow(
            survey_number=str(r.survey_number),
            recommendation=r.recommendation,
            area_ha=area,
            chain_coverage=float(getattr(r, "chain_coverage", 0.0) or 0.0),
            field_residual_max=float(getattr(r, "field_residual_max", float("inf"))),
            match_method=getattr(r, "match_method", "") or "",
        ))
    rows.sort(key=lambda x: (x.recommendation not in _ACCEPT, x.survey_number))
    return rows


def _disposition_counts(rows: list[VillageRow]) -> dict[str, int]:
    out: dict[str, int] = {}
    for r in rows:
        out[r.recommendation] = out.get(r.recommendation, 0) + 1
    return out


def village_area_statement_pdf(rows: list[VillageRow], village: str = "",
                               crs: str = "") -> bytes:
    """Render the village area-statement PDF; returns bytes."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    buf = io.BytesIO()
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm,
                            topMargin=18 * mm, bottomMargin=18 * mm)
    title = ParagraphStyle("T", parent=styles["Heading1"], fontSize=14,
                           alignment=TA_CENTER, spaceAfter=2)
    sub = ParagraphStyle("S", parent=styles["Normal"], fontSize=9,
                         alignment=TA_CENTER, textColor=colors.grey)
    elements: list[Any] = [
        Paragraph("Survey and Settlement Department", sub),
        Paragraph("Government of Tamil Nadu", sub),
        Spacer(1, 3 * mm),
        Paragraph("VILLAGE GEOREFERENCING STATEMENT", title),
        Paragraph(f"{village or '—'}    CRS {crs or '—'}", sub),
        Spacer(1, 4 * mm),
    ]

    header = ["Survey No.", "Disposition", "Area (Ha)", "Chain Cover", "Field Resid (m)", "Method"]
    data: list[list[Any]] = [header]
    style_rows: list[tuple] = []
    accept_bg = colors.HexColor("#E6F4EA")
    review_bg = colors.HexColor("#FFF2CC")
    nocov_bg = colors.HexColor("#F0F0F0")
    for i, r in enumerate(rows, start=1):
        fr = "—" if r.field_residual_max != r.field_residual_max or r.field_residual_max == float("inf") \
            else f"{r.field_residual_max:.2f}"
        data.append([
            r.survey_number,
            r.recommendation.replace("_", " ").title(),
            f"{r.area_ha:.4f}" if r.area_ha is not None else "—",
            f"{100 * r.chain_coverage:.0f}%" if r.georeferenced or r.chain_coverage else "—",
            fr,
            r.match_method or "—",
        ])
        bg = accept_bg if r.georeferenced else review_bg if r.recommendation == "REVIEW" else nocov_bg
        style_rows.append(("BACKGROUND", (0, i), (-1, i), bg))

    table = Table(data, colWidths=[24 * mm, 30 * mm, 24 * mm, 26 * mm, 30 * mm, 28 * mm],
                  repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9D9D9")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ] + style_rows))
    elements.append(table)
    elements.append(Spacer(1, 4 * mm))

    disp = _disposition_counts(rows)
    placed = sum(v for k, v in disp.items() if k in _ACCEPT)
    total_area = sum(r.area_ha for r in rows if r.area_ha is not None)
    summary = ParagraphStyle("Sum", parent=styles["Normal"], fontSize=9)
    elements.append(Paragraph(
        f"<b>Summary:</b> {len(rows)} survey(s) — {placed} georeferenced, "
        + ", ".join(f"{v} {k.replace('_', ' ').title()}"
                    for k, v in sorted(disp.items()) if k not in _ACCEPT)
        + f". Total georeferenced area: {total_area:.4f} ha.", summary))
    elements.append(Spacer(1, 2 * mm))
    elements.append(Paragraph(
        "<i>Georeferenced plots are placed rigidly (rotation + scale~1 + translation) "
        "from field/cadastral control; geometry is the FMB vector boundary, never warped. "
        "REVIEW = needs operator confirmation; No Coverage = not field-traced.</i>",
        ParagraphStyle("N", parent=styles["Normal"], fontSize=7, textColor=colors.grey)))

    doc.build(elements)
    return buf.getvalue()


def village_excel(rows: list[VillageRow]) -> bytes:
    """Render the village Excel breakdown; returns bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Village Georeferencing"
    headers = ["Survey No", "Disposition", "Georeferenced", "Area (Ha)",
               "Chain Coverage", "Field Residual (m)", "Match Method"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor="D9D9D9")
    for r in rows:
        ws.append([
            r.survey_number,
            r.recommendation.replace("_", " ").title(),
            "Yes" if r.georeferenced else "No",
            round(r.area_ha, 4) if r.area_ha is not None else None,
            round(r.chain_coverage, 3),
            (None if r.field_residual_max in (float("inf"),) or r.field_residual_max != r.field_residual_max
             else round(r.field_residual_max, 3)),
            r.match_method or "",
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for i, w in enumerate([12, 16, 14, 12, 14, 18, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_village_delivery(
    results: Iterable[Any],
    combined_dxf: str | Path | None,
    output_dir: str | Path,
    village: str = "",
    crs: str = "",
) -> Path:
    """Write the village PDF + Excel and bundle them with the combined DXF into a zip.

    Returns the path to the delivery zip. Best-effort and local-only (no S3
    dependency) so it can run at the end of every corridor job; the API/Job path
    keeps using ``package.package_and_deliver`` for S3 upload.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    rows = rows_from_results(results)

    pdf_bytes = village_area_statement_pdf(rows, village=village, crs=crs)
    xlsx_bytes = village_excel(rows)
    (output_dir / "village_area_statement.pdf").write_bytes(pdf_bytes)
    (output_dir / "village_area_breakdown.xlsx").write_bytes(xlsx_bytes)

    zip_path = output_dir / "village_delivery.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("village_area_statement.pdf", pdf_bytes)
        zf.writestr("village_area_breakdown.xlsx", xlsx_bytes)
        if combined_dxf is not None and Path(combined_dxf).exists():
            zf.write(combined_dxf, arcname=Path(combined_dxf).name)
    _log.info("Village delivery package: %s (%d surveys)", zip_path, len(rows))
    return zip_path
