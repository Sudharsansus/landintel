"""M4 report tests: generate real PDF and Excel from a real M1 plot, verify content.

Uses survey 100 from the full M1 chain (OCR + anchor + build + validate + anomaly)
to produce the actual artefacts, then asserts:
- PDF is valid (starts with %PDF), has non-trivial size, and can be re-read.
- Excel has the right sheets, the right header columns, and the plot's data in it.
- Package zips both artefacts and the DXF, uploads to moto-mocked S3, returns a
  presigned URL.

Real-fixture suite so OCR runs once (module-scoped fixture). Report the collected
count proactively alongside pass/fail.
"""

from __future__ import annotations

import io
import zipfile
from pathlib import Path

import moto
import boto3
import openpyxl
import pytest

from landintel.agent.anomaly import check_plot
from landintel.agent.validator import validate_plot
from landintel.config import get_settings
from landintel.core.enums import PlotStatus
from landintel.core.models import Job, Plot
from landintel.pipeline.m1_extract.anchor import anchor_measurements
from landintel.pipeline.m1_extract.build_plot import build_plot
from landintel.pipeline.m1_extract.ocr import extract_text, parse_header
from landintel.pipeline.m1_extract.pdf_vectors import extract_vectors
from landintel.pipeline.m1_extract.to_dxf import write_dxf
from landintel.pipeline.m4_report.area_statement import generate_area_statement
from landintel.pipeline.m4_report.excel_sheet import generate_excel_sheet
from landintel.pipeline.m4_report.package import package_and_deliver

FMB_DIR = Path(__file__).resolve().parents[1] / "fixtures" / "FMB"
BUCKET = "landintel-test"
REGION = "ap-south-1"


@pytest.fixture(scope="module")
def real_job(tmp_path_factory) -> Job:
    """Build survey 100 through M1 + validate + anomaly; return a Job with one plot."""
    f = FMB_DIR / "FMB_SIVAGANGAI_Manamadurai_T.Pudukkottai_100.pdf"
    vectors = extract_vectors(f)
    detections = extract_text(f)
    header = parse_header(detections)
    plot = build_plot(
        client_id="client_test",
        vectors=vectors,
        detections=detections,
        anchor_result=anchor_measurements(vectors, detections),
        header=header,
    )
    validate_plot(plot, client=None)
    check_plot(plot)

    out = tmp_path_factory.mktemp("m4_out")
    dxf_path = write_dxf(plot, out / "survey_100.dxf")

    job = Job(client_id="client_test", input_files=[str(f)],
              output_files=[str(dxf_path)])
    job.plots.append(plot)
    return job


# --- PDF area statement ------------------------------------------------------


def test_pdf_is_valid_bytes(real_job: Job) -> None:
    pdf = generate_area_statement(real_job)
    assert isinstance(pdf, bytes) and len(pdf) > 2_000
    assert pdf[:4] == b"%PDF", "does not start with PDF magic"


def test_pdf_contains_survey_number(real_job: Job, tmp_path: Path) -> None:
    """The PDF text layer contains the survey number (smoke-read via pypdf)."""
    pytest.importorskip("pypdf")
    from pypdf import PdfReader
    pdf = generate_area_statement(real_job)
    reader = PdfReader(io.BytesIO(pdf))
    text = " ".join(page.extract_text() or "" for page in reader.pages)
    assert "100" in text
    assert "Sivagangai" in text


def test_pdf_area_values_present(real_job: Job, tmp_path: Path) -> None:
    """The stated and computed areas appear in the PDF text."""
    pytest.importorskip("pypdf")
    from pypdf import PdfReader
    pdf = generate_area_statement(real_job)
    text = " ".join(
        p.extract_text() or "" for p in PdfReader(io.BytesIO(pdf)).pages
    )
    # stated area 1.665 and computed ~1.697 should both appear
    assert "1.665" in text or "1.667" in text or "1.697" in text or "1.70" in text


# --- Excel area sheet --------------------------------------------------------


def test_excel_is_valid_xlsx(real_job: Job) -> None:
    xlsx = generate_excel_sheet(real_job)
    assert isinstance(xlsx, bytes) and len(xlsx) > 1_000
    assert xlsx[:4] == b"PK\x03\x04", "does not start with xlsx/zip magic"


def test_excel_has_correct_sheets(real_job: Job) -> None:
    xlsx = generate_excel_sheet(real_job)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    assert "Area Summary" in wb.sheetnames
    assert "Measurements" in wb.sheetnames


def test_excel_area_summary_has_plot_row(real_job: Job) -> None:
    xlsx = generate_excel_sheet(real_job)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb["Area Summary"]
    # Row 1 is the header; row 2 should be survey 100.
    assert ws.cell(row=1, column=1).value == "Survey No."
    assert ws.cell(row=2, column=1).value == "100"
    # Stated area is in column 5.
    stated = ws.cell(row=2, column=5).value
    assert stated is not None and abs(float(stated) - 1.665) < 0.001


def test_excel_measurements_sheet_has_accepted(real_job: Job) -> None:
    xlsx = generate_excel_sheet(real_job)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb["Measurements"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # Survey 100 has ~27 accepted measurements.
    assert len(rows) >= 10, f"expected >=10 accepted measurements, got {len(rows)}"
    # Every row's survey_no is "100".
    assert all(r[0] == "100" for r in rows if r[0] is not None)


# --- Package + S3 delivery ---------------------------------------------------


@pytest.fixture
def s3_env(monkeypatch):
    monkeypatch.setenv("AWS_ACCESS_KEY_ID", "test-key")
    monkeypatch.setenv("AWS_SECRET_ACCESS_KEY", "test-secret")
    monkeypatch.setenv("S3_BUCKET", BUCKET)
    monkeypatch.setenv("ANTHROPIC_API_KEY", "sk-ant-test")
    get_settings.cache_clear()
    yield
    get_settings.cache_clear()


def test_package_returns_presigned_url(real_job: Job, s3_env) -> None:
    with moto.mock_aws():
        boto3.client("s3", region_name=REGION).create_bucket(
            Bucket=BUCKET,
            CreateBucketConfiguration={"LocationConstraint": REGION},
        )
        url = package_and_deliver(real_job)
    assert isinstance(url, str) and url.startswith("https://")
    assert "delivery" in url or "landintel" in url


def test_package_zip_contains_expected_files(real_job: Job, s3_env) -> None:
    with moto.mock_aws():
        s3 = boto3.client("s3", region_name=REGION)
        s3.create_bucket(
            Bucket=BUCKET,
            CreateBucketConfiguration={"LocationConstraint": REGION},
        )
        package_and_deliver(real_job)
        # Fetch the uploaded object and inspect the zip.
        key = f"client_test/jobs/{real_job.id}/landintel_delivery.zip"
        obj = s3.get_object(Bucket=BUCKET, Key=key)
        zip_bytes = obj["Body"].read()

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
    assert "area_statement.pdf" in names
    assert "area_breakdown.xlsx" in names
    # The DXF should also be in the zip.
    assert any(n.endswith(".dxf") for n in names), f"no dxf in zip: {names}"
